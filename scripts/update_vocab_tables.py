"""Update vocabulary tables in md content files.

Task 1: fix 20260326-final.md (two 4-col tables -> 5-col with phonetic placeholder)
Task 2: enrich listening md vocab tables with Example Sentences from xlsx
"""
from __future__ import annotations

import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]


def clean_example(s: str | None) -> str:
    if s is None:
        return "-"
    text = str(s).strip()
    if not text:
        return "-"
    if text in ("(no sentence)", "—"):
        return "-"
    # pipes break markdown rows
    text = text.replace("|", "/")
    return text


def load_single_sheet(path: Path) -> dict[str, str]:
    """Sheet layout: row1=header, rows after=data with (#, Word, Example)."""
    wb = openpyxl.load_workbook(path)
    ws = wb["Vocabulary"]
    out: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        if not isinstance(row[0], (int, float)):
            continue
        word = str(row[1]).strip().lower() if row[1] else ""
        ex = clean_example(row[2])
        if word:
            out[word] = ex
    return out


def load_dual_sheet(path: Path) -> tuple[dict[str, str], dict[str, str]]:
    """Layout:
    - 'Academic Vocabulary' title row, then header (#, Word, Example Sentence, Src)
    - blank row
    - 'Domain Vocabulary' title row, then header (#, Term, Definition, Example Sentence, Src)
    Returns (academic_map, domain_map)
    """
    wb = openpyxl.load_workbook(path)
    ws = wb["Vocabulary"]
    rows = list(ws.iter_rows(values_only=True))

    academic: dict[str, str] = {}
    domain: dict[str, str] = {}
    mode: str | None = None  # 'academic' | 'domain'

    for row in rows:
        if not row:
            continue
        c0 = row[0]
        # section titles
        if c0 == "Academic Vocabulary":
            mode = "academic"
            continue
        if c0 == "Domain Vocabulary":
            mode = "domain"
            continue
        # header row has '#'
        if c0 == "#":
            continue
        # numbered data rows
        if not isinstance(c0, (int, float)):
            continue
        if mode == "academic":
            word = str(row[1]).strip().lower() if row[1] else ""
            ex = clean_example(row[2])
            if word:
                academic[word] = ex
        elif mode == "domain":
            # Term in col 1, Example in col 3 (0-indexed), Definition in col 2
            word = str(row[1]).strip().lower() if row[1] else ""
            ex = clean_example(row[3]) if len(row) > 3 else "-"
            if word:
                domain[word] = ex
    return academic, domain


def update_table(
    lines: list[str],
    start_idx: int,
    lookup: dict[str, str],
) -> tuple[int, int]:
    """Update the vocab table starting at header line `start_idx`.
    Returns (rows_updated, rows_missed).
    Mutates lines in place.
    """
    # Header
    assert lines[start_idx].strip() == "| 單字 | 音標 | 詞性 | 中文 |"
    lines[start_idx] = "| 單字 | 音標 | 詞性 | 中文 | 例句 |"
    # Separator
    sep_idx = start_idx + 1
    assert lines[sep_idx].strip().startswith("|------|------|------|------|")
    lines[sep_idx] = "|------|------|------|------|------|"

    updated = 0
    missed = 0
    i = sep_idx + 1
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        if not stripped.startswith("|"):
            break
        # parse existing cells
        cells = [c.strip() for c in stripped.split("|")]
        # cells[0] empty (leading |), cells[-1] empty (trailing |)
        if len(cells) < 6:
            # malformed; stop
            break
        word_cell = cells[1]
        key = word_cell.strip().lower()
        example = lookup.get(key, "-")
        if example == "-":
            missed += 1
        else:
            updated += 1
        new_line = f"| {cells[1]} | {cells[2]} | {cells[3]} | {cells[4]} | {example} |"
        # preserve original trailing whitespace/newline style
        lines[i] = new_line
        i += 1
    return updated, missed


def process_file(
    md_path: Path,
    lookups: list[dict[str, str]],
) -> list[tuple[int, int]]:
    """Process a md file where lookups are applied to tables in order of appearance.

    Returns list of (updated, missed) per table processed.
    """
    text = md_path.read_text(encoding="utf-8")
    # preserve exact newline handling
    # split into lines without losing trailing newline
    if text.endswith("\n"):
        has_trailing = True
        body = text[:-1]
    else:
        has_trailing = False
        body = text
    lines = body.split("\n")

    # find header rows (4-col form)
    header_str = "| 單字 | 音標 | 詞性 | 中文 |"
    header_indices = [i for i, ln in enumerate(lines) if ln.strip() == header_str]

    assert len(header_indices) == len(lookups), (
        f"{md_path}: found {len(header_indices)} tables, expected {len(lookups)}"
    )

    results = []
    for hi, lu in zip(header_indices, lookups):
        results.append(update_table(lines, hi, lu))

    new_text = "\n".join(lines) + ("\n" if has_trailing else "")
    md_path.write_text(new_text, encoding="utf-8")
    return results


def process_task1() -> None:
    """Fix 20260326-final.md: 4-col (word, pos, meaning, synonym) -> 5-col (word, -, pos, meaning, synonym)."""
    md = ROOT / "src/content/reading/toefl/20260326-final.md"
    text = md.read_text(encoding="utf-8")
    if text.endswith("\n"):
        has_trailing = True
        body = text[:-1]
    else:
        has_trailing = False
        body = text
    lines = body.split("\n")

    header_str = "| 單字 | 詞性 | 中文 | 同義詞 |"
    header_indices = [i for i, ln in enumerate(lines) if ln.strip() == header_str]
    assert len(header_indices) == 2, f"Expected 2 tables in final.md, got {len(header_indices)}"

    total_rows = 0
    for hi in header_indices:
        lines[hi] = "| 單字 | 音標 | 詞性 | 中文 | 同義詞 |"
        sep_idx = hi + 1
        assert lines[sep_idx].strip().startswith("|------|------|------|--------|")
        lines[sep_idx] = "|------|------|------|------|--------|"
        i = sep_idx + 1
        while i < len(lines):
            stripped = lines[i].strip()
            if not stripped.startswith("|"):
                break
            cells = [c.strip() for c in stripped.split("|")]
            if len(cells) < 6:
                break
            # cells: ['', word, pos, meaning, synonym, '']
            word, pos, meaning, synonym = cells[1], cells[2], cells[3], cells[4]
            lines[i] = f"| {word} | - | {pos} | {meaning} | {synonym} |"
            total_rows += 1
            i += 1

    md.write_text("\n".join(lines) + ("\n" if has_trailing else ""), encoding="utf-8")
    print(f"Task 1: 20260326-final.md — {total_rows} rows column-shifted across 2 tables")


def process_task2() -> None:
    # sp2 art history
    art_map = load_single_sheet(
        ROOT / "course_files/Listening_Speaking/114-2-spring/sp2-AH-art_history/sp2_vocab.xlsx"
    )
    art_res = process_file(
        ROOT / "src/content/listening/toefl/20260315-art-history.md",
        [art_map],
    )
    for idx, (u, m) in enumerate(art_res, 1):
        print(f"  20260315-art-history.md table {idx}: {u} matched, {m} fallback(-)")

    # sp3 science
    sci_map = load_single_sheet(
        ROOT / "course_files/Listening_Speaking/114-2-spring/sp3/sp3_vocab.xlsx"
    )
    sci_res = process_file(
        ROOT / "src/content/listening/toefl/20260322-science.md",
        [sci_map],
    )
    for idx, (u, m) in enumerate(sci_res, 1):
        print(f"  20260322-science.md table {idx}: {u} matched, {m} fallback(-)")

    # sp4 geology
    geo_map = load_single_sheet(
        ROOT / "course_files/Listening_Speaking/114-2-spring/sp4/sp4_vocab.xlsx"
    )
    geo_res = process_file(
        ROOT / "src/content/listening/toefl/20260329-geology.md",
        [geo_map],
    )
    for idx, (u, m) in enumerate(geo_res, 1):
        print(f"  20260329-geology.md table {idx}: {u} matched, {m} fallback(-)")

    # sp05 business (dual)
    biz_ac, biz_dm = load_dual_sheet(
        ROOT / "course_files/Listening_Speaking/114-2-spring/sp05-SS-Business/sp05_vocab.xlsx"
    )
    biz_res = process_file(
        ROOT / "src/content/listening/toefl/20260412-business.md",
        [biz_ac, biz_dm],
    )
    print(f"  20260412-business.md academic: {biz_res[0][0]} matched, {biz_res[0][1]} fallback(-)")
    print(f"  20260412-business.md domain:   {biz_res[1][0]} matched, {biz_res[1][1]} fallback(-)")

    # sp06 zoology (dual)
    zoo_ac, zoo_dm = load_dual_sheet(
        ROOT / "course_files/Listening_Speaking/114-2-spring/sp06-LS-Zoology/sp06_vocab.xlsx"
    )
    zoo_res = process_file(
        ROOT / "src/content/listening/toefl/20260419-zoology.md",
        [zoo_ac, zoo_dm],
    )
    print(f"  20260419-zoology.md academic: {zoo_res[0][0]} matched, {zoo_res[0][1]} fallback(-)")
    print(f"  20260419-zoology.md domain:   {zoo_res[1][0]} matched, {zoo_res[1][1]} fallback(-)")


if __name__ == "__main__":
    process_task1()
    process_task2()
